#! python3
#multiplicationTable.py -Create a program multiplicationTable.py 
#that takes a number N from the command line and creates an N×N 
#multiplication table in an Excel spreadsheet.

import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()

sheet = wb['Sheet']

bold_font = Font(bold = True)

#receive user input
size = input('How big should the table be? \n')
size = int(size)

#set up the labels
for i in range (1,size+1):
	row = sheet.cell(row=i+1, column = 1)
	row.value = i
	row.font = bold_font
	column = sheet.cell(row=1, column = i+1)
	column.value = i
	column.font = bold_font

#create the multiplication table
for j in range (1, size+1):
	for i in range (1,size+1):
		column = sheet.cell(row = 1+j, column = i+1)
		column.value = i*j

	


wb.save('multiplicationTable.xls')